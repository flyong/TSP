# -*- coding: utf-8 -*-
import os, sys
import numpy as np
import pywt
from ts_benchmark.data.dataprocessor.getTimeRange import read_excel_data
import pandas as pd
from multiprocessing import Pool
from statsmodels.tsa.stattools import grangercausalitytests
from ts_benchmark.data.dataprocessor.plottest import plotNetwork
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import ts_benchmark.data.dataprocessor.extra as ex
from ts_benchmark.common.constant import ROOT_PATH
from ts_benchmark.data.dataprocessor.getTimeRange import get_value_range_by_date_range
from ts_benchmark.common.constant import CONFIG_PATH
import json


class HiddenPrints:
    """
    class to hide the print output
    """

    def __enter__(self):
        self._original_stdout = sys.stdout
        sys.stdout = open(os.devnull, "w")

    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stdout.close()
        sys.stdout = self._original_stdout


# fuction to realise Wavelet transformation
def wavelet_transform(data, wavelet_name="db2"):
    # perform wavelet transformation
    cA, cD = pywt.dwt(data, wavelet_name)
    return cA, cD


# function to realise wavelet decomposition
def wavelet_decomposition(data, wavelet_name="db2", wavelet_level=4):
    # perform wavelet decomposition
    ca = []
    for i in range(wavelet_level):
        cA = pywt.downcoef("a", data, wavelet_name, level=i + 1)
        ca.append(cA)
    cd = pywt.wavedec(data, wavelet_name, level=wavelet_level)
    cd = cd[1:]
    # merge the lists ca and cd
    result = ca + cd
    return result


def correlation_coefficient(ts1, ts2):
    """
    function to calculate the correlation coefficient between two time series
    note the time series should have the same dimension
    ts1: time series 1
    ts2: time series 2
    """
    cor_list = []
    for i in range(len(ts1)):
        cor_list.append(
            np.corrcoef(ts1[i], ts2[i])[0, 1]
        )  # correlation coefficient between cA, cD
    return cor_list


# function to calculate the correlation coefficient between two time series
def cross_correlation(ground_settlement_WF, tunnel_settlement_WF):
    correlation_GS_TS = []
    for ca in ground_settlement_WF:
        for cb in tunnel_settlement_WF:
            cor = correlation_coefficient(ca[1], cb[1])
            # insert the groundID and tunnelID to the correlation list
            cor.insert(0, ca[0])
            cor.insert(1, cb[0])
            correlation_GS_TS.append(cor)
    return correlation_GS_TS


# return the list of sensors in the convergent section
def getSensorList_conver():
    groundSensorID = [
        "GS10-1",
        "GS10-2",
        "GS10-3",
        "GS10-4",
        "GS10-5",
        "GS10-6",
        "GS10-7",
        "GS10-8",
        "GS10-9",
        "GS8-1",
        "GS8-10",
        "GS8-2",
        "GS8-3",
        "GS8-5",
        "GS8-6",
        "GS8-7",
        "GS8-8",
        "GS8-9",
        "GS9-1",
        "GS9-10",
        "GS9-2",
        "GS9-5",
        "GS9-6",
        "GS9-7",
        "GS9-8",
        "GS9-9",
    ]

    tunnelSensorID = [
        "TCA1-01",
        "TCA1-02",
        "TCA1-03",
        "TCA1-04",
        "TCA1-05",
        "TCA1-06",
        "TCA1-07",
        "TCA1-08",
        "TCA1-09",
        "TCA1-10",
        "TCA1-11",
        "TCA1-12",
        "TCA1-13",
        "TCA1-14",
        "TCA1-15",
        "TCA1-16",
        "TCA1-17",
        "TCA1-18",
        "TCA1-19",
        "TCA1-20",
        "TCA1-21",
        "TCA1-22",
        "TCA1-23",
        "TCA1-24",
        "TCA1-25",
        "TCA1-26",
        "TCA1-27",
        "TCA1-28",
        "TCA1-29",
        "TCA1-30",
        "TCA1-31",
        "TCA1-32",
        "TCA1-33",
        "TCA1-34",
        "TCA1-35",
        "TCA1-36",
        "TCA1-37",
        "TCA1-38",
        "TCA1-39",
        "TCA1-40",
        "TCA1-41",
        "TCA1-42",
        "TCA1-43",
        "TCA1-44",
        "TCA1-45",
        "TCA1-46",
        "TCA1-47",
        "TCA1-48",
        "TCA1-49",
        "TCA1-50",
        "TCA1-51",
        "TCA1-52",
        "TCA1-53",
        "TCA1-54",
        "TCA1-55",
        "TCA1-56",
        "TCA1-57",
        "TCA1-58",
        "TCA1-59",
        "TCA1-60",
        "TCA1-61",
        "TCA1-62",
    ]  # Sensors in tunnelA1


# function to definte time range
def get_date_range() -> tuple:
    """
    return the start date and end date of the data"""
    with open(os.path.join(CONFIG_PATH, "common_config.json"), "r") as file:
        common_config = json.load(file)
    start_date = common_config["start_date"]
    end_date = common_config["end_date"]
    return start_date, end_date


# function to calculate the correlation coefficient within time series
def self_correlation(ground_settlement_WF):
    self_correlation = []
    for ca in ground_settlement_WF:
        for cb in ground_settlement_WF:
            cor = correlation_coefficient(ca[1], cb[1])
            # insert the groundID and tunnelID to the correlation list
            cor.insert(0, ca[0])
            cor.insert(1, cb[0])
            self_correlation.append(cor)
    return self_correlation


# funtion to select data from the same time range
def select_data_by_time_range(data, start_time, end_time):
    selected_data = data[(data["Date"] >= start_time) & (data["Date"] <= end_time)]
    return selected_data


# function to write correlation matrix
def writeDataToExcel(wtresults, file_path, sheetname):
    # transpose wtresults
    wtresults = np.array(wtresults).T
    final_df = pd.DataFrame(wtresults, columns=["ca1", "cd1"])
    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name=sheetname, index=False)
    print(f"WTresults has been successfully written to sheet in {file_path}")


# function to write correlation matrix to excel
def writeCorrelationToExcel(correlation_GS_TS, file_path, sheetname):
    # convert the list of list to a dataframe and write to excel
    final_df = pd.DataFrame(
        correlation_GS_TS,
        columns=["groundID", "tunnelID", "1", "2", "3", "4", "5", "6", "7", "8"],
    )
    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name=sheetname, index=False)
    print(f"Correlation has been successfully written to sheet in {file_path}")


def runWT():
    """
    function to get wavelet transformation and correlation coefficient
    """
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    sheet_name = [
        "GroundSettlement",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
        "A1ConverM",
    ]

    grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
    grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])

    start_time, end_time = get_date_range()

    # wavelet transformation parameters

    wavelet_level = 4
    wavelet_name = "db2"

    ground_settlement_WF = []
    for groundID in groundSensorID:
        sensor_data = grouped_groundsettle.get_group(groundID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        ground_settlement_WF.append([groundID, wtResult])

    # release memory of grouped data
    grouped_groundsettle = None

    tunnel_settlement_WF = []
    for tunnelID in tunnelSensorID:
        sensor_data = grouped_tunnelsettle.get_group(tunnelID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        tunnel_settlement_WF.append([tunnelID, wtResult])

    # release memory of grouped data
    grouped_tunnelsettle = None
    self_correlation_GS = self_correlation(ground_settlement_WF)
    writeCorrelationToExcel(self_correlation_GS, correlation_path, "selfcorrelationGS")

    # calculate correlation coefficient between ground settlement and tunnel settlement
    # correlation_GS_TS= cross_correlation(ground_settlement_WF, tunnel_settlement_WF)
    # writeCorrelationToExcel(correlation_GS_TS,correlation_path,'correlationTSGS')


def WT_correlation_by_date_range(
    start_time, end_time, grouped_groundsettle=None, grouped_tunnelsettle=None
) -> pd.DataFrame:
    """
    rewrite function to get wavelet transformation and correlation coefficient
    start_time: str, start time of the data
    end_time: str, end time of the data
    grouped_groundsettle: optional, pre-grouped ground settlement data
    grouped_tunnelsettle: optional, pre-grouped tunnel settlement data
    """
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    sheet_name = [
        "GroundSettlement",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
        "A1ConverM",
    ]

    if grouped_groundsettle is None:
        grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
        grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])

    # wavelet transformation parameters

    wavelet_level = 4
    wavelet_name = "db2"

    ground_settlement_WF = []
    for groundID in groundSensorID:
        sensor_data = grouped_groundsettle.get_group(groundID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        ground_settlement_WF.append([groundID, wtResult])

    # release memory of grouped data
    grouped_groundsettle = None

    tunnel_settlement_WF = []
    for tunnelID in tunnelSensorID:
        sensor_data = grouped_tunnelsettle.get_group(tunnelID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        tunnel_settlement_WF.append([tunnelID, wtResult])

    # release memory of grouped data
    grouped_tunnelsettle = None

    correlation_GS_TS = cross_correlation(ground_settlement_WF, tunnel_settlement_WF)
    final_df = pd.DataFrame(
        correlation_GS_TS,
        columns=["groundID", "tunnelID", "1", "2", "3", "4", "5", "6", "7", "8"],
    )

    return final_df


# calculate the correlation coefficient between ground settlement and tunnel settlement
def correlateSPC():
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.ex.getSensorList()
    start_time, end_time = get_date_range()

    sheet_name = [
        "GroundSettlement",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
    ]

    grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
    grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])

    # wavelet transformation parameters

    wavelet_level = 4
    wavelet_name = "db2"

    ground_settlement_WF = []
    for groundID in groundSensorID:
        sensor_data = grouped_groundsettle.get_group(groundID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        ground_settlement_WF.append([groundID, wtResult])

    # release memory of grouped data
    grouped_groundsettle = None

    tunnel_settlement_WF = []
    for tunnelID in tunnelSensorID:
        sensor_data = grouped_tunnelsettle.get_group(tunnelID)
        selected_data = select_data_by_time_range(sensor_data, start_time, end_time)
        wtResult = wavelet_decomposition(
            selected_data["value"], wavelet_name, wavelet_level
        )
        tunnel_settlement_WF.append([tunnelID, wtResult])

    # Spearman's rank correlation coefficient
    spearman_correlation_GS_TS = []
    for ca in ground_settlement_WF:
        for cb in tunnel_settlement_WF:
            cor = correlation_coefficient(ca[1], cb[1])
            # insert the groundID and tunnelID to the correlation list
            cor.insert(0, ca[0])
            cor.insert(1, cb[0])
            spearman_correlation_GS_TS.append(cor)


# function to generate correlation matrix
def generateMatrix():
    file_path = r"/home/vsc/workspace/IDT3/Cor/dataprocess/data/correlation.xlsx"
    sheet_name_list = [
        "Self-GS-level1",
        "selfcorrelationGS",
        "correlation",
        "selfcorrelationGS",
    ]

    sheet_name = sheet_name_list[2]
    level = "1"

    excel_file = pd.ExcelFile(file_path)
    excel_data = excel_file
    df = excel_data.parse(sheet_name)
    # Transform a, b, relation value list to a relation matrix
    relation_matrix = df.pivot_table(
        index="tunnelID", columns="groundID", values="CR" + level
    )

    # write the relation matrix to excel
    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
        relation_matrix.to_excel(
            writer, sheet_name=sheet_name + "Matrix" + level, index=True
        )
    print(f"Relation matrix has been successfully written to sheet in {file_path}")


def read_excel(file_path, sheet_name):
    # read the data from the excel file
    excel_file = pd.ExcelFile(file_path)
    excel_data = excel_file
    df = excel_data.parse(sheet_name)
    return df


def append_df_to_excel(file_path, sheet_name, dataframe):
    """
    Append a pandas DataFrame to an existing Excel sheet.

    :param file_path: str, path to the Excel file
    :param sheet_name: str, name of the sheet to append data to
    :param dataframe: pandas DataFrame containing the data to append
    """
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)

    # Check if the sheet exists, if not create it
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    sheet = workbook[sheet_name]

    # Get the next empty row
    next_row = sheet.max_row + 1

    # Write the column headers if the sheet is empty
    if next_row == 1:
        for col, header in enumerate(dataframe.columns, start=1):
            sheet.cell(row=1, column=col, value=header)
        next_row = 2

    # Append the data
    for row in dataframe.itertuples(index=False):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=next_row, column=col, value=value)
        next_row += 1

    # Save the workbook
    workbook.save(file_path)
    print(f"Data appended successfully to {file_path}")


# function to calculate the distance between GS and TS
def calculateDistance():

    position_path = r"/home/vsc/workspace/IDT3/Cor/dataprocess/data/position.xlsx"
    df_position = read_excel(position_path, "GS_XY")
    df_position_TS = read_excel(position_path, "TS_A1_M_XY")

    # calculate the distance between GS and TS
    distance = []
    for row1 in df_position.iterrows():
        # get the x and y position of GS
        x1 = row1[1]["x_position"]
        y1 = row1[1]["y_position"]

        for row2 in df_position_TS.iterrows():
            x2 = row2[1]["x_position"]
            y2 = row2[1]["y_position"]
            d = np.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)
            distance.append([row1[1]["sensorID"], row2[1]["sensorID"], d])

    distance = pd.DataFrame(distance, columns=["sensoraID", "sensorbID", "distance"])
    distanceMatrix = distance.pivot_table(
        index="sensoraID", columns="sensorbID", values="distance"
    )

    # write the distance matrix to excel
    with pd.ExcelWriter(position_path, mode="a", engine="openpyxl") as writer:
        distanceMatrix.to_excel(writer, sheet_name="distanceMatrix", index=True)
    print(f"Distance matrix has been successfully written to sheet in {position_path}")


# function to calculate Granger Causality between two time series
def grangerCausality(ts1, ts2):
    # import the package for Granger Causality
    data = np.array([ts1, ts2]).T
    # perform Granger Causality test
    result = grangercausalitytests(data, maxlag=5)
    return result


# function to calculate correlation matrix
def calculateCorrelationMatrix(
    threshold,
    pearsonThreshold,
    distanceThreshold,
    maxConnection,
    minValue_GS,
    minValue_TS,
    start_time,
    end_time,
    writeList,
    overwrite_sheet,
):
    # read the correlation matrix
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    correlation_sheet = "correlation"
    range_sheet = [
        "GroundSettlementInterpolated",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
    ]
    sheet_name = [
        "GroundSettlement",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
    ]

    df_rawRelation = read_excel(correlation_path, correlation_sheet)

    grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
    grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])
    groundData_rangeList = read_excel(range_path, range_sheet[0])
    tunnelData_rangeList = read_excel(range_path, range_sheet[3])

    df_distance = read_excel(position_path, "distanceMatrix")

    df_relation = pd.DataFrame(columns=["groundID", "tunnelID", "GCT", "relation"])
    for groundID in groundSensorID:
        groundData = select_data_by_time_range(
            grouped_groundsettle.get_group(groundID), start_time, end_time
        )
        groundData_range = groundData_rangeList.loc[
            (groundData_rangeList["sensorID"] == groundID), "min_value"
        ].values[0]

        for tunnelID in tunnelSensorID:

            tunnelData = select_data_by_time_range(
                grouped_tunnelsettle.get_group(tunnelID), start_time, end_time
            )
            tunnelData_range = tunnelData_rangeList.loc[
                (tunnelData_rangeList["sensorID"] == tunnelID), "min_value"
            ].values[0]

            # get the distance between GS and TS
            distance = df_distance.loc[
                (df_distance["sensoraID"] == groundID), tunnelID
            ].values[0]

            # get the average Pearson correlation coefficient
            averagePearson = (
                df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "CR1",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "CR2",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "CR3",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "CR4",
                ].values[0]
            ) / 4

            # calculate the Granger Causality
            correlation = grangerCausality(groundData["value"], tunnelData["value"])
            ssr_ftest = correlation[1][0]["ssr_ftest"][1]

            # check the relation between GS and TS
            if (
                ssr_ftest < threshold
                and distance < distanceThreshold
                and averagePearson > pearsonThreshold
                and groundData_range < minValue_GS
                and tunnelData_range < minValue_TS
            ):
                df_relation = pd.concat(
                    [
                        df_relation,
                        pd.DataFrame(
                            [
                                {
                                    "groundID": groundID,
                                    "tunnelID": tunnelID,
                                    "GCT": ssr_ftest,
                                    "relation": 1,
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
            else:
                df_relation = pd.concat(
                    [
                        df_relation,
                        pd.DataFrame(
                            [
                                {
                                    "groundID": groundID,
                                    "tunnelID": tunnelID,
                                    "GCT": ssr_ftest,
                                    "relation": 0,
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )

    # write the relation matrix to excel
    if writeList:
        with pd.ExcelWriter(
            correlation_path, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            df_relation.to_excel(writer, sheet_name="relationList", index=False)
        print(
            f"Relation matrix has been successfully written to sheet in {correlation_path}"
        )

    # write the relation matrix to excel
    pivot_table = df_relation.pivot_table(
        index="tunnelID", columns="groundID", values="relation"
    )
    if overwrite_sheet != "":
        with pd.ExcelWriter(
            correlation_path, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            pivot_table.to_excel(writer, sheet_name=overwrite_sheet, index=True)
        print(
            f"Relation matrix has been successfully written to sheet in {correlation_path}"
        )
    else:
        with pd.ExcelWriter(
            correlation_path, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            pivot_table.to_excel(writer, sheet_name="relationMatrix", index=True)
        print(
            f"Relation matrix has been successfully written to sheet in {correlation_path}"
        )


# function to test parameters
def testParameters(
    thresholdlist,
    pearsonThreshold,
    distanceThreshold,
    maxConnection,
    minValue_GS,
    minValue_TS,
    gcRelation,
    figureIndex=00,
):
    """
    test the parameters of the correlation matrix without writing the result to excel
    """
    # read the correlation matrix
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    correlation_sheet = "correlation"
    range_sheet = [
        "GroundSettlementInterpolated",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
    ]

    df_rawRelation = read_excel(correlation_path, correlation_sheet)

    groundData_rangeList = read_excel(range_path, range_sheet[0])
    tunnelData_rangeList = read_excel(range_path, range_sheet[3])

    df_distance = read_excel(position_path, "distanceMatrix")

    for threshold in thresholdlist:
        df_relation = pd.DataFrame(columns=["groundID", "tunnelID", "GCT", "relation"])
        for groundID in groundSensorID:

            groundData_range = groundData_rangeList.loc[
                (groundData_rangeList["sensorID"] == groundID), "min_value"
            ].values[0]

            for tunnelID in tunnelSensorID:

                tunnelData_range = tunnelData_rangeList.loc[
                    (tunnelData_rangeList["sensorID"] == tunnelID), "min_value"
                ].values[0]

                # get the distance between GS and TS
                distance = df_distance.loc[
                    (df_distance["sensoraID"] == groundID), tunnelID
                ].values[0]

                # get the average Pearson correlation coefficient
                averagePearson = (
                    df_rawRelation.loc[
                        (df_rawRelation["groundID"] == groundID)
                        & (df_rawRelation["tunnelID"] == tunnelID),
                        "CR1",
                    ].values[0]
                    + df_rawRelation.loc[
                        (df_rawRelation["groundID"] == groundID)
                        & (df_rawRelation["tunnelID"] == tunnelID),
                        "CR2",
                    ].values[0]
                    + df_rawRelation.loc[
                        (df_rawRelation["groundID"] == groundID)
                        & (df_rawRelation["tunnelID"] == tunnelID),
                        "CR3",
                    ].values[0]
                    + df_rawRelation.loc[
                        (df_rawRelation["groundID"] == groundID)
                        & (df_rawRelation["tunnelID"] == tunnelID),
                        "CR4",
                    ].values[0]
                ) / 4

                # calculate the Granger Causality

                ssr_ftest = gcRelation.loc[
                    (gcRelation["groundID"] == groundID)
                    & (gcRelation["tunnelID"] == tunnelID),
                    "grrelation",
                ].values[0]

                # check the relation between GS and TS
                if (
                    ssr_ftest < threshold
                    and distance < distanceThreshold
                    and averagePearson > pearsonThreshold
                    and groundData_range < minValue_GS
                    and tunnelData_range < minValue_TS
                ):
                    df_relation = pd.concat(
                        [
                            df_relation,
                            pd.DataFrame(
                                [
                                    {
                                        "groundID": groundID,
                                        "tunnelID": tunnelID,
                                        "GCT": ssr_ftest,
                                        "relation": 1,
                                    }
                                ]
                            ),
                        ],
                        ignore_index=True,
                    )
                else:
                    df_relation = pd.concat(
                        [
                            df_relation,
                            pd.DataFrame(
                                [
                                    {
                                        "groundID": groundID,
                                        "tunnelID": tunnelID,
                                        "GCT": ssr_ftest,
                                        "relation": 0,
                                    }
                                ]
                            ),
                        ],
                        ignore_index=True,
                    )
        NetWorkFeature = analyseNetwork2(df_relation)
        NetWorkFeature["threshold"] = threshold
        NetWorkFeature["pearsonThreshold"] = pearsonThreshold
        NetWorkFeature["distanceThreshold"] = distanceThreshold
        NetWorkFeature["maxConnection"] = maxConnection
        NetWorkFeature["minValue_GS"] = minValue_GS
        NetWorkFeature["minValue_TS"] = minValue_TS
        NetWorkFeature["figureIndex"] = figureIndex
        writeNetworkFeatureToExcel(pd.DataFrame([NetWorkFeature]), "netfeature")


# function to test parameters
def rolling_calculate_correlation(
    start_date, end_date, write_flag=False, test_flag=False
):
    """
    calculate the correlation matrix with rolling window
    """

    # pre-defined parameters
    threshold = 0.05
    pearsonThreshold = 0.5
    maxConnection = 5
    minValue_GS = -5
    minValue_TS = -5
    distanceThreshold = 100
    sheet_name_correlation = "relationMatrix_rolling"

    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    sheet_name = ex.get_sheet_name_list("interpolated.xlsx")

    # avoid repeated reading of the data
    grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
    grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])

    # get value range in the specific time range
    groundData_rangeList = get_value_range_by_date_range(
        grouped_groundsettle, start_date, end_date
    )
    tunnelData_rangeList = get_value_range_by_date_range(
        grouped_tunnelsettle, start_date, end_date
    )

    # calculate the pearson correlation matrix
    print("Calculating the Pearson correlation matrix...")
    df_rawRelation = WT_correlation_by_date_range(
        start_date, end_date, grouped_groundsettle, grouped_tunnelsettle
    )

    # calculate the Granger Causality matrix
    print(start_date, "to ", end_date, " calculating the Granger Causality matrix...")
    with HiddenPrints():
        gcRelation = calculate_GcMatrix(
            start_date, end_date, grouped_groundsettle, grouped_tunnelsettle
        )

    df_distance = read_excel(position_path, "distanceMatrix")

    df_relation = pd.DataFrame(columns=["groundID", "tunnelID", "GCT", "relation"])
    for groundID in groundSensorID:

        groundData_range = groundData_rangeList.loc[
            (groundData_rangeList["sensorID"] == groundID), "min_value"
        ].values[0]

        for tunnelID in tunnelSensorID:

            tunnelData_range = tunnelData_rangeList.loc[
                (tunnelData_rangeList["sensorID"] == tunnelID), "min_value"
            ].values[0]

            # get the distance between GS and TS
            distance = df_distance.loc[
                (df_distance["sensoraID"] == groundID), tunnelID
            ].values[0]

            # get the average Pearson correlation coefficient
            averagePearson = (
                df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "1",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "2",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "3",
                ].values[0]
                + df_rawRelation.loc[
                    (df_rawRelation["groundID"] == groundID)
                    & (df_rawRelation["tunnelID"] == tunnelID),
                    "4",
                ].values[0]
            ) / 4

            # calculate the Granger Causality

            ssr_ftest = gcRelation.loc[
                (gcRelation["groundID"] == groundID)
                & (gcRelation["tunnelID"] == tunnelID),
                "grrelation",
            ].values[0]

            # check the relation between GS and TS
            if (
                ssr_ftest < threshold
                and distance < distanceThreshold
                and abs(averagePearson) > pearsonThreshold
                and groundData_range < minValue_GS
                and tunnelData_range < minValue_TS
            ):
                df_relation = pd.concat(
                    [
                        df_relation,
                        pd.DataFrame(
                            [
                                {
                                    "groundID": groundID,
                                    "tunnelID": tunnelID,
                                    "GCT": ssr_ftest,
                                    "relation": 1,
                                    "distance": distance,
                                    "averagePearson": averagePearson,
                                    "groundData_range": groundData_range,
                                    "tunnelData_range": tunnelData_range,
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
            else:
                df_relation = pd.concat(
                    [
                        df_relation,
                        pd.DataFrame(
                            [
                                {
                                    "groundID": groundID,
                                    "tunnelID": tunnelID,
                                    "GCT": ssr_ftest,
                                    "relation": 0,
                                    "distance": distance,
                                    "averagePearson": averagePearson,
                                    "groundData_range": groundData_range,
                                    "tunnelData_range": tunnelData_range,
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
    if write_flag:
        pivot_table = df_relation.pivot_table(
            index="tunnelID", columns="groundID", values="relation"
        )
        with pd.ExcelWriter(
            correlation_path, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            pivot_table.to_excel(writer, sheet_name=sheet_name_correlation, index=True)

    # test_flag = True
    if test_flag:
        test_path = os.path.join(ROOT_PATH, "dataset", "test.xlsx")
        with pd.ExcelWriter(
            test_path, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            df_relation.to_excel(writer, sheet_name="relationList", index=False)
        print(f"Relation matrix has been successfully written to sheet in {test_path}")

    with open(os.path.join(CONFIG_PATH, "common_config.json"), "r") as file:
        common_config = json.load(file)

    figureIndex = common_config["figure_index"]
    NetWorkFeature = analyseNetwork2(df_relation)
    NetWorkFeature.update(
        {
            "threshold": threshold,
            "pearsonThreshold": pearsonThreshold,
            "distanceThreshold": distanceThreshold,
            "maxConnection": maxConnection,
            "minValue_GS": minValue_GS,
            "minValue_TS": minValue_TS,
            "figureIndex": figureIndex,
            "start_date": start_date,
            "end_date": end_date,
        }
    )

    writeNetworkFeatureToExcel(pd.DataFrame([NetWorkFeature]), "netfeature")

    plotNetwork(figureIndex, sheet_name_correlation, df_relation)
    figureIndex += 1

    # write the figure index to the common config file
    with open(os.path.join(CONFIG_PATH, "common_config.json"), "w") as file:
        common_config["figure_index"] = figureIndex
        json.dump(common_config, file)


# calculate gr correlation matrix
def calculate_GcMatrix(
    start_time, end_time, grouped_groundsettle=None, grouped_tunnelsettle=None
):
    # read the correlation matrix
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )

    groundSensorID, tunnelSensorID = ex.getSensorList()

    sheet_name = [
        "GroundSettlement",
        "B1SettleM",
        "B2SettleM",
        "A1SettleM",
        "A2SettleM",
    ]

    if grouped_groundsettle is None:
        grouped_groundsettle = read_excel_data(data_path, sheet_name[0])
    if grouped_tunnelsettle is None:
        grouped_tunnelsettle = read_excel_data(data_path, sheet_name[3])

    df_grrelation = pd.DataFrame(columns=["groundID", "tunnelID", "grrelation"])

    for groundID in groundSensorID:
        groundData = select_data_by_time_range(
            grouped_groundsettle.get_group(groundID), start_time, end_time
        )

        for tunnelID in tunnelSensorID:

            tunnelData = select_data_by_time_range(
                grouped_tunnelsettle.get_group(tunnelID), start_time, end_time
            )

            # calculate the Granger Causality
            correlation = grangerCausality(groundData["value"], tunnelData["value"])
            ssr_ftest = correlation[1][0]["ssr_ftest"][1]
            df_grrelation = pd.concat(
                [
                    df_grrelation,
                    pd.DataFrame(
                        [
                            {
                                "groundID": groundID,
                                "tunnelID": tunnelID,
                                "grrelation": ssr_ftest,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )

    return df_grrelation


# analyse network topology and features
def analyseNetwork(data_sheet="relationMatrix2"):

    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    groundSensorID, tunnelSensorID = ex.getSensorList()

    sheet_name = "netfeature"

    df_relation = read_excel(correlation_path, data_sheet)

    # calculate average degree of ground sensors
    linkedGroundCount = 0
    totalGroundDegree = 0
    maxGroundDegree = 0
    totalDegree = 0
    for groundID in groundSensorID:
        linkedCount = 0
        for tunnelID in tunnelSensorID:
            if (
                df_relation.loc[(df_relation["tunnelID"] == tunnelID), groundID].values[
                    0
                ]
                == 1
            ):
                linkedCount += 1
                totalGroundDegree += 1
        if linkedCount > 0:
            linkedGroundCount += 1
        if linkedCount > maxGroundDegree:
            maxGroundDegree = linkedCount
    averageGroundDegree = totalGroundDegree / linkedGroundCount

    # calculate average degree of tunnel sensors
    linkedTunnelCount = 0
    totalTunnelDegree = 0
    maxTunnelDegree = 0
    for tunnelID in tunnelSensorID:
        linkedCount = 0
        for groundID in groundSensorID:
            if (
                df_relation.loc[(df_relation["tunnelID"] == tunnelID), groundID].values[
                    0
                ]
                == 1
            ):
                linkedCount += 1
                totalTunnelDegree += 1
        if linkedCount > 0:
            linkedTunnelCount += 1
        if linkedCount > maxTunnelDegree:
            maxTunnelDegree = linkedCount
    averageTunnelDegree = totalTunnelDegree / linkedTunnelCount

    # calculate average degree of the network discard the isolated nodes
    averageDegree = (totalGroundDegree + totalTunnelDegree) / (
        linkedGroundCount + linkedTunnelCount
    )
    clusterCoefficient = (totalGroundDegree + totalTunnelDegree) / (
        len(groundSensorID) * len(tunnelSensorID)
    )

    result = {
        "averageGroundDegree": averageGroundDegree,
        "maxGroundDegree": maxGroundDegree,
        "linkedGroundCount": linkedGroundCount,
        "averageTunnelDegree": averageTunnelDegree,
        "maxTunnelDegree": maxTunnelDegree,
        "linkedTunnelCount": linkedTunnelCount,
        "averageDegree": averageDegree,
        "clusterCoefficient": clusterCoefficient,
    }
    return result


# directly analyse the network topology and features without reading from excel
def analyseNetwork2(df_relation) -> dict:
    """
    analyse the network topology and features without reading from excel
    df_relation: pd.DataFrame, relation matrix

    """

    groundSensorID, tunnelSensorID = ex.getSensorList()
    # calculate average degree of ground sensors
    linkedGroundCount = 0
    totalGroundDegree = 0
    maxGroundDegree = 0
    totalDegree = 0
    for groundID in groundSensorID:
        linkedCount = 0
        for tunnelID in tunnelSensorID:
            relation = df_relation.loc[
                (df_relation["tunnelID"] == tunnelID)
                & (df_relation["groundID"] == groundID),
                "relation",
            ].values[0]
            if relation == 1:
                linkedCount += 1
                totalGroundDegree += 1
        if linkedCount > 0:
            linkedGroundCount += 1
        if linkedCount > maxGroundDegree:
            maxGroundDegree = linkedCount
    averageGroundDegree = totalGroundDegree / linkedGroundCount

    # calculate average degree of tunnel sensors
    linkedTunnelCount = 0
    totalTunnelDegree = 0
    maxTunnelDegree = 0
    for tunnelID in tunnelSensorID:
        linkedCount = 0
        for groundID in groundSensorID:
            relation = df_relation.loc[
                (df_relation["tunnelID"] == tunnelID)
                & (df_relation["groundID"] == groundID),
                "relation",
            ].values[0]
            if relation:
                linkedCount += 1
                totalTunnelDegree += 1
        if linkedCount > 0:
            linkedTunnelCount += 1
        if linkedCount > maxTunnelDegree:
            maxTunnelDegree = linkedCount
    averageTunnelDegree = totalTunnelDegree / linkedTunnelCount

    # calculate average degree of the network discard the isolated nodes
    averageDegree = (totalGroundDegree + totalTunnelDegree) / (
        linkedGroundCount + linkedTunnelCount
    )
    clusterCoefficient = (totalGroundDegree + totalTunnelDegree) / (
        len(groundSensorID) * len(tunnelSensorID)
    )

    result = {
        "averageGroundDegree": averageGroundDegree,
        "maxGroundDegree": maxGroundDegree,
        "linkedGroundCount": linkedGroundCount,
        "averageTunnelDegree": averageTunnelDegree,
        "maxTunnelDegree": maxTunnelDegree,
        "linkedTunnelCount": linkedTunnelCount,
        "averageDegree": averageDegree,
        "clusterCoefficient": clusterCoefficient,
    }
    return result


def writeNetworkFeatureToExcel(result, sheet_name):
    data_path, result_path, correlation_path, position_path, range_path = (
        ex.getFilePathList()
    )
    # write the relation matrix to excel
    append_df_to_excel(correlation_path, sheet_name, result)
    print(
        f"Network features has been successfully written to sheet in {correlation_path}"
    )


# function to calculate the relation matrix
def calculateRelationMatrix():
    threshold = 0.08  # threshold for Granger Causality
    pearsonThreshold = 0.8
    distanceThreshold = 300
    maxConnection = 5
    minValue_GS = -10
    minValue_TS = -5
    start_time = "2020-11-2"
    end_time = "2021-12-11"
    writeList = False
    sheet_name = "relationMatrix2"
    figureIndex = 21
    for j in range(1, 6):
        if j == 1:
            distanceThreshold = 200
        elif j == 2:
            distanceThreshold = 150
        elif j == 3:
            distanceThreshold = 100
        elif j == 4:
            distanceThreshold = 70
        elif j == 5:
            distanceThreshold = 50
        for i in range(1, 4):
            if i == 1:
                threshold = 0.08
            elif i == 2:
                threshold = 0.06
            elif i == 3:
                threshold = 0.05

            print(f"Threshold: {threshold}, DistanceThreshold: {distanceThreshold}")

            calculateCorrelationMatrix(
                threshold,
                pearsonThreshold,
                distanceThreshold,
                maxConnection,
                minValue_GS,
                minValue_TS,
                start_time,
                end_time,
                writeList,
                sheet_name,
            )

            result = analyseNetwork(sheet_name)

            # add threshold to the result
            result["threshold"] = threshold
            result["pearsonThreshold"] = pearsonThreshold
            result["distanceThreshold"] = distanceThreshold
            result["maxConnection"] = maxConnection
            result["minValue_GS"] = minValue_GS
            result["minValue_TS"] = minValue_TS
            result["figureIndex"] = figureIndex

            writeNetworkFeatureToExcel(pd.DataFrame([result]), "netfeature")
            plotNetwork(figureIndex, sheet_name)
            figureIndex += 1


def run_parameter_Test():
    """
    test the parameters for the correlation analysis
    """
    threshold = 0.05  # threshold for Granger Causality
    pearsonThreshold = 0.8
    distanceThreshold = 100
    maxConnection = 5
    minValue_GS = 0
    minValue_TS = 0
    start_time = "2020-11-2"
    end_time = "2021-12-11"
    writeList = True
    sheet_name = "relationMatrix2"
    figureIndex = 00

    # calculateCorrelationMatrix(threshold,pearsonThreshold,distanceThreshold,maxConnection,minValue_GS,minValue_TS,start_time,end_time,writeList,sheet_name)
    # result = analyseNetwork(sheet_name)
    df_grrelation = calculate_GcMatrix(start_time, end_time)
    thresholdList = []
    for i in range(11):
        threshold = 0.04 + i * 0.01
        thresholdList.append(threshold)
    testParameters(
        thresholdList,
        pearsonThreshold,
        distanceThreshold,
        maxConnection,
        minValue_GS,
        minValue_TS,
        df_grrelation,
        figureIndex,
    )

    # plotNetwork(figureIndex,sheet_name)


def analyseCorrelation():
    """
    run the selected correlation analysis
    """
    threshold = 0.05  # threshold for Granger Causality
    pearsonThreshold = 0.8
    distanceThreshold = 100
    maxConnection = 5
    minValue_GS = -5
    minValue_TS = -5
    start_time = "2020-11-2"
    end_time = "2021-12-11"
    writeList = True
    sheet_name = "relationMatrix2"
    figureIndex = 31

    # calculateCorrelationMatrix(threshold,pearsonThreshold,distanceThreshold,maxConnection,minValue_GS,minValue_TS,start_time,end_time,writeList,sheet_name)
    # result = analyseNetwork(sheet_name)

    calculateCorrelationMatrix(
        threshold,
        pearsonThreshold,
        distanceThreshold,
        maxConnection,
        minValue_GS,
        minValue_TS,
        start_time,
        end_time,
        writeList,
        sheet_name,
    )

    result = analyseNetwork(sheet_name)
    # add threshold to the result
    result["threshold"] = threshold
    result["pearsonThreshold"] = pearsonThreshold
    result["distanceThreshold"] = distanceThreshold
    result["maxConnection"] = maxConnection
    result["minValue_GS"] = minValue_GS
    result["minValue_TS"] = minValue_TS
    result["figureIndex"] = figureIndex
    writeNetworkFeatureToExcel(pd.DataFrame([result]), "netfeature")
    # plotNetwork(figureIndex,sheet_name)


# runWT()

# generateMatrix()

# calculateRelationMatrix()

# plotNetwork('relationMatrix2')

# analyseCorrelation()

# runparameterTest()

# TODO: rolling time range based correlation analysis: finished
# 1. calculate the correlation matrix for each time range: finished
# 2. reconduct the correlation analysis for the whole network: finished
